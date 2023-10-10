import win32com.client
import pandas as pd
from datetime import datetime
import subprocess
import re
import os
import sys
import time
from openpyxl import load_workbook, Workbook
import pyperclip

# Librerias a instalar despues de instalar Python
'''
    pip install pandas
    pip install pywin32
    
'''

# Conexiones para poder entrar a SAP GUI
try:
    sapGuiAuto = win32com.client.GetObject('SAPGUI')
    application = sapGuiAuto.GetScriptingEngine
    connection = application.Children(0)
    session = connection.Children(0)
except Exception as e:
    print(f"Recuerda primero abrir SAP en la transacción Z2L, vuelve a correr el comando")
    time.sleep(5)
    input("Presiona Enter para continuar.")
    sys.exit(1)


now = datetime.now()
dt_string = now.strftime("%d_%m_%Y_%H_%M")

# Obtener la ruta del escritorio
desktop = os.path.join(os.path.expanduser('~'), 'OneDrive - BASF' , 'Desktop')
username = os.getlogin()

# Definir el nombre de la carpeta
nombreCarpeta = "SAP_Automatizacion"

# Nombre del material
nombre = ""

# Unir la ruta del escritorio con el nombre de la carpeta
folderdir = os.path.join(desktop, nombreCarpeta)

# Crear la carpeta
try:
    os.mkdir(folderdir)
except OSError as e:
    print(f"No se pudo crear la carpeta '{nombreCarpeta}': {e}")

# Fechas
lastYear = now.year - 1
startDate=f'{now.day}.{now.month}.{lastYear}'
endDate=f'{now.day}.{now.month}.{now.year}'

def main():
    print('''

 /$$$$$$$  /$$                                                   /$$       /$$          
| $$__  $$|__/                                                  |__/      | $$          
| $$  \ $$ /$$  /$$$$$$  /$$$$$$$  /$$    /$$ /$$$$$$  /$$$$$$$  /$$  /$$$$$$$  /$$$$$$ 
| $$$$$$$ | $$ /$$__  $$| $$__  $$|  $$  /$$//$$__  $$| $$__  $$| $$ /$$__  $$ /$$__  $$
| $$__  $$| $$| $$$$$$$$| $$  \ $$ \  $$/$$/| $$$$$$$$| $$  \ $$| $$| $$  | $$| $$  \ $$
| $$  \ $$| $$| $$_____/| $$  | $$  \  $$$/ | $$_____/| $$  | $$| $$| $$  | $$| $$  | $$
| $$$$$$$/| $$|  $$$$$$$| $$  | $$   \  $/  |  $$$$$$$| $$  | $$| $$|  $$$$$$$|  $$$$$$/
|_______/ |__/ \_______/|__/  |__/    \_/    \_______/|__/  |__/|__/ \_______/ \______/ 
                                                                                        
                                                                                        
                                                                                        

''')
    print(f"Hola {username}, excelente día")

    numerosSAP = []
    while(True):
        opc = input("Escribe el o los codigos SAP a sacar datos (1. Terminar): ")
        if opc == "1":
            break
        elif re.match(r'^\d{8}$', opc):
            numerosSAP.append(opc)
        else:
            print("No es un número SAP, intente de nuevo")
    
    if len(numerosSAP) > 0:
        print(f"Los codigos introducidos son: {numerosSAP}")
        sapLoader(numerosSAP)


def sapLoader(numerosSAP):
    try:
        # Ingreso a transacción YGQM_ILC
        # Diccionario de codigos SAP con su respectiva ruta a Excel
        excels = {}
        
        try:
            for codigo in numerosSAP:
                filename = "Lotes_SAP_" + codigo + "_" + dt_string + ".xlsx"
                excels[codigo] = filename

                session.findById("wnd[0]").maximize()
                session.findById("wnd[0]/tbar[0]/okcd").text = "YGQM_ILC"
                session.findById("wnd[0]").sendVKey(0)
                session.findById("wnd[0]/usr/tabsTABSTRIP_TBLOCK/tabpTAB1/ssub%_SUBSCREEN_TBLOCK:/BASF/YGQM_ILC:0001/ctxtS_WERKS-LOW").text = "MX31"
                session.findById("wnd[0]/usr/tabsTABSTRIP_TBLOCK/tabpTAB1/ssub%_SUBSCREEN_TBLOCK:/BASF/YGQM_ILC:0001/ctxtSO_PASTR-LOW").text = startDate
                session.findById("wnd[0]/usr/tabsTABSTRIP_TBLOCK/tabpTAB1/ssub%_SUBSCREEN_TBLOCK:/BASF/YGQM_ILC:0001/ctxtSO_PASTR-HIGH").text = endDate
                session.findById("wnd[0]/usr/tabsTABSTRIP_TBLOCK/tabpTAB1/ssub%_SUBSCREEN_TBLOCK:/BASF/YGQM_ILC:0001/ctxtSO_MATNR-LOW").text = codigo
                session.findById("wnd[0]/usr/tabsTABSTRIP_TBLOCK/tabpTAB1/ssub%_SUBSCREEN_TBLOCK:/BASF/YGQM_ILC:0001/ctxtSO_MATNR-LOW").setFocus
                session.findById("wnd[0]/usr/tabsTABSTRIP_TBLOCK/tabpTAB1/ssub%_SUBSCREEN_TBLOCK:/BASF/YGQM_ILC:0001/ctxtSO_MATNR-LOW").caretPosition = 8


                session.findById("wnd[0]/tbar[1]/btn[8]").press()
                session.findById("wnd[0]/shellcont/shell").pressToolbarButton ("&MB_VARIANT")
                session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").currentCellRow = -1
                session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectColumn ("SELTEXT")
                session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").contextMenu()
                total_rows = session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").rowCount
                session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = f"0-{total_rows - 1}"

                session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/btnAPP_WL_SING").press()
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
                session.findById("wnd[0]/shellcont/shell").pressToolbarContextButton("&MB_EXPORT")
                session.findById("wnd[0]/shellcont/shell").selectContextMenuItem("&XXL")
                session.findById("wnd[1]/usr/ctxtDY_PATH").text = folderdir
                session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = filename
                session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 10
                session.findById("wnd[1]/tbar[0]/btn[11]").press()
                session.findById("wnd[0]/tbar[0]/btn[12]").press()
                session.findById("wnd[0]/tbar[0]/btn[12]").press()

            # Todos los codigos SAP exportados, cerrar Excel
            os.system("taskkill /f /im EXCEL.exe")

        except AttributeError:
            # os.system("taskkill /f /im saplogon.exe")
            print("Error, puede que algun codigo SAP no exista")

        time.sleep(5)
        
        # Iterar diccionarios de excel para sacar batches y ponerlo en su respectivo excel
        for clave, nombreArchivo in excels.items():
            
            rutaCompleta = folderdir + "\\" + nombreArchivo

            # Cargar el archivo excel en un DataFrame de pandas 
            # Aqui las columnas seleccionadas son con base a lo establecido previamente, se pueden cambiar a voluntad
            df = pd.read_excel(rutaCompleta, header=0, usecols=lambda x: re.search("Mat|text|Batch|batch|Visc|visc|resist|electrical|Electrical|viscosidad|pH|ph|PH|lote real|Lote Real|lote Real|Actual lot quantity|Actual Lot Quantity|actual lot quantity|Insp. start date|insp. Start Date", x), sheet_name="Sheet1")

            # // regex
            # /visc(?:osidad|osity|)\s*,?\s*(?:shear)?\s*(?:\[\d*\])?/i

            # Convertir la columna de fechas a objetos datetime de Python
            df["Insp. start date"] = pd.to_datetime(df["Insp. start date"], errors='coerce')

            # Dar formato a las fechas en el estilo "dd.mm.yyyy"
            df["Insp. start date"] = df["Insp. start date"].dt.strftime('%d.%m.%Y')

            # Ordenar el DataFrame por la columna 'Start date' en orden ascendente (más antiguo al más nuevo)
            df["Insp. start date"] = pd.to_datetime(df["Insp. start date"], format='%d.%m.%Y')
            df = df.sort_values("Insp. start date", ascending=True)

            # Volver a convertir la columna de fechas a su formato original
            df["Insp. start date"] = df["Insp. start date"].dt.strftime('%d.%m.%Y')


            # Guardar el Excel
            df.to_excel(rutaCompleta, sheet_name="YGQM_ILC", columns=[ x for x in df.columns if re.search("Mat|text|Batch|batch|Visc|visc|resist|electrical|Electrical|viscosidad|pH|ph|PH|lote real|Lote Real|lote Real|Actual lot quantity|Actual Lot Quantity|actual lot quantity|Insp. start date|insp. Start Date", x)], index=False)

            # Obtener nombre
            nombre = df["Short text for inspection object"][0]

            # Operaciones
            batches = df["Batch"]
            batches = [str(lote) for lote in batches ]


            # Quitar los batches que tengan mas de 6 ceros
            batches = [element for element in batches if element.count("0") <= 4]

            # Revisar si no tiene un .n extra o no es un numero
            batches = [str(lote).split(".")[0] for lote in batches if not pd.isnull(lote) and lote != 'nan']

            # Cambiar nombre
            nuevoNombre = str(nombre) + " - " + str(clave) + "_" + str(dt_string) + ".xlsx"
            nuevaRuta = folderdir + "\\" + nuevoNombre
            os.rename(rutaCompleta, nuevaRuta)

            time.sleep(1)
            
            #df = pd.read_excel(nuevaRuta)
            
            # Diccionario de lotes con información
            lotesConInfo = {}
            
            try:
                # Recorrer lotes con YC096 para sacar info
                for lote in batches:
                    session.findById("wnd[0]").maximize()
                    session.findById("wnd[0]/tbar[0]/okcd").text = "YCO96"
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[0]/usr/ctxtP_WERKS").text = "MX31"
                    session.findById("wnd[0]/usr/ctxtS_AUFNR-LOW").text = lote
                    session.findById("wnd[0]/usr/ctxtS_AUFNR-LOW").setFocus()
                    session.findById("wnd[0]/usr/ctxtS_AUFNR-LOW").caretPosition = 9
                    session.findById("wnd[0]/tbar[1]/btn[8]").press()
                    session.findById("wnd[0]/mbar/menu[0]/menu[1]/menu[2]").select()
                    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[4,0]").select()
                    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[4,0]").setFocus()
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    session.findById("wnd[0]/tbar[0]/btn[12]").press()
                    session.findById("wnd[0]/tbar[0]/btn[12]").press()

                    # Guardar el texto copiado por SAP en el diccionario
                    lotesConInfo[lote] = pyperclip.paste()
            
            except Exception as e:
                print(f"Error, es posible que : {lote} no tenga registro")

            # Leer excel, poner hoja nueva de YCO96 e ir poniendo info
            libro = load_workbook(nuevaRuta)

            # Verificar si la hoja "YCO96" ya existe en el libro y eliminarla
            if "YCO96" in libro.sheetnames:
                libro.remove(libro["YCO96"])

            # Crear un nuevo objeto worksheet con la hoja nueva
            worksheet = libro.create_sheet("YCO96")

            # Recorrer el diccionario y escribir los valores en la columna B
            fila = 1  # Iniciar la fila en 1
            for clave, valor in lotesConInfo.items():
                # Escribir la clave en la celda A de la fila correspondiente
                worksheet.cell(row=fila, column=1, value=clave)
                # Dividir el valor en líneas usando el carácter de salto de línea "\n"
                lineas = valor.split("\n")
                # Recorrer cada línea del valor y escribirla en las celdas a partir de la columna B de la fila correspondiente
                for linea in lineas:
                    # Dividir la línea en palabras usando espacios
                    palabras = linea.split()
                    columna = 2  # Iniciar la columna en 2 (columna B)

                    # Recorrer cada palabra en la lista de palabras
                    for i, palabra in enumerate(palabras):
                        # # Verificar si la siguiente palabra contiene "KG", "kg" o "%" y fusionarla con la palabra actual
                        # if i < len(palabras) - 1 and any(x in palabras[i+1] for x in ["KG", "kg", "%", "K", "k"]):
                        #     palabra += " " + palabras[i+1]
                        #     palabras.pop(i+1)  # Eliminar la siguiente palabra de la lista, ya que se fusionó con la palabra actual

                        # Escribir la palabra en la celda correspondiente
                        worksheet.cell(row=fila, column=columna, value=palabra)
                        columna += 1  # Incrementar la columna en 1

                    fila += 1  # Incrementar la fila en 1

            # Guardar el archivo modificado
            libro.save(nuevaRuta)

        # Correr la sesión de SAP
        connection = None
        application = None
        sapGuiAuto = None
    except Exception as e:
        print(f"No se pudo completar, el error es: {e}")
        input("Presione enter para cerrar")
        

if __name__ == "__main__":
    main()



'''
Wiki botones
- Boton amarillo de cancelar - session.findById("wnd[0]/tbar[0]/btn[15]").press()
- Boton rojo de cerrar - session.findById("wnd[0]/tbar[0]/btn[12]").press()

'''